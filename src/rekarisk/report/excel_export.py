"""
Rekarisk Excel Export.

Multi-sheet Excel workbook export for consequence analysis results.
Sheets:
    1. Summary — all scenarios with key results
    2. Input Parameters — complete input data per scenario
    3. Results Tables — detailed distance-based results
    4. Risk Results — IR values, FN data, risk matrix
    5. Comparison — side-by-side scenario comparison
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Union

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, NamedStyle,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference, LineChart
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ══════════════════════════════════════════════════════════════════════════════
# Style Constants
# ══════════════════════════════════════════════════════════════════════════════

_HEADER_FILL = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_SUBHEADER_FILL = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
_SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_DATA_FONT = Font(name="Calibri", size=10)
_TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1A5276")
_ALT_FILL = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin", color="AEB6BF"),
    right=Side(style="thin", color="AEB6BF"),
    top=Side(style="thin", color="AEB6BF"),
    bottom=Side(style="thin", color="AEB6BF"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _apply_header_style(ws, row: int, cols: int, fill=_HEADER_FILL, font=_HEADER_FONT):
    """Apply header styling to a row."""
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER


def _apply_data_style(ws, start_row: int, end_row: int, cols: int, alignment=_LEFT):
    """Apply data cell styling with alternating row colors."""
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = _DATA_FONT
            cell.alignment = alignment
            cell.border = _THIN_BORDER
            if (r - start_row) % 2 == 1:
                cell.fill = _ALT_FILL


def _auto_width(ws, min_width: int = 10, max_width: int = 40):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split("\n")
                cell_len = max(len(line) for line in lines)
                max_len = max(max_len, cell_len + 2)
        ws.column_dimensions[col_letter].width = min(max_len, max_width)


# ══════════════════════════════════════════════════════════════════════════════
# Public API
# ══════════════════════════════════════════════════════════════════════════════

def export_to_excel(
    project_data: Dict[str, Any],
    results: List[Dict[str, Any]],
    output_path: Union[str, Path],
) -> str:
    """Export project data and results to a multi-sheet Excel workbook.

    Parameters
    ----------
    project_data : dict
        Project metadata (name, description, scenarios, weather_cases, etc.).
    results : list[dict]
        List of result dicts, one per scenario.
    output_path : str or Path
        Where to save the .xlsx file.

    Returns
    -------
    str
        The output path.
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install with: pip install openpyxl"
        )

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    _build_summary_sheet(wb, project_data, results)
    _build_input_sheet(wb, project_data, results)
    _build_results_sheet(wb, results)
    _build_risk_sheet(wb, results)
    _build_comparison_sheet(wb, results)

    # Remove default empty sheet if any
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 5:
        del wb["Sheet"]

    wb.save(str(output_path))
    return str(output_path)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet Builders
# ══════════════════════════════════════════════════════════════════════════════

def _build_summary_sheet(
    wb: Workbook,
    project_data: Dict[str, Any],
    results: List[Dict[str, Any]],
) -> None:
    """Sheet 1: Summary."""
    ws = wb.active
    ws.title = "Summary"

    # Title
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"Rekarisk Consequence Analysis — {project_data.get('name', 'Untitled')}"
    title_cell.font = _TITLE_FONT
    title_cell.alignment = _CENTER

    # Project info
    ws.merge_cells("A3:B3")
    ws["A3"].value = "Project Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True, color="1A5276")

    info = [
        ("Project Name", project_data.get("name", "")),
        ("Description", project_data.get("description", "")),
        ("Created", project_data.get("created_at", "")),
        ("Version", project_data.get("format_version", "")),
        ("Scenarios", len(results)),
    ]
    for i, (label, value) in enumerate(info):
        row = 4 + i
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=row, column=2, value=str(value)).font = _DATA_FONT

    # Scenarios summary table
    table_start = 10
    ws.merge_cells(f"A{table_start}:E{table_start}")
    ws.cell(row=table_start, column=1, value="Scenario Summary").font = Font(
        name="Calibri", size=12, bold=True, color="1A5276"
    )

    headers = ["#", "Name", "Type", "Substance", "Key Result"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=table_start + 1, column=c, value=h)
    _apply_header_style(ws, table_start + 1, len(headers))

    for i, res in enumerate(results, 1):
        row = table_start + 1 + i
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=res.get("name", ""))
        ws.cell(row=row, column=3, value=res.get("type", ""))
        ws.cell(row=row, column=4, value=_extract_substance(res))
        # Key result
        summary = res.get("summary", {})
        key = summary.get("max_concentration") or summary.get("max_radiation") or \
              summary.get("max_overpressure") or summary.get("mass_released") or ""
        ws.cell(row=row, column=5, value=str(key))

    _apply_data_style(ws, table_start + 2, table_start + 1 + len(results), len(headers))
    _auto_width(ws)


def _build_input_sheet(
    wb: Workbook,
    project_data: Dict[str, Any],
    results: List[Dict[str, Any]],
) -> None:
    """Sheet 2: Input Parameters."""
    ws = wb.create_sheet("Input Parameters")

    ws.merge_cells("A1:D1")
    ws["A1"].value = "Input Parameters"
    ws["A1"].font = _TITLE_FONT

    row = 3

    # Weather cases
    weather_cases = project_data.get("weather_cases", [])
    if weather_cases:
        ws.merge_cells(f"A{row}:D{row}")
        ws.cell(row=row, column=1, value="Weather Cases").font = Font(
            name="Calibri", size=12, bold=True, color="1A5276"
        )
        row += 1

        w_headers = ["Name", "Wind Speed (m/s)", "Direction (°)", "Stability", "Temperature (°C)", "Humidity (%)"]
        for c, h in enumerate(w_headers, 1):
            ws.cell(row=row, column=c, value=h)
        _apply_header_style(ws, row, len(w_headers))
        row += 1

        for w in weather_cases:
            ws.cell(row=row, column=1, value=w.get("name", ""))
            ws.cell(row=row, column=2, value=w.get("wind_speed", ""))
            ws.cell(row=row, column=3, value=w.get("wind_direction", ""))
            ws.cell(row=row, column=4, value=w.get("stability_class", ""))
            ws.cell(row=row, column=5, value=w.get("temperature", w.get("ambient_temp", "")))
            ws.cell(row=row, column=6, value=w.get("humidity", ""))
            row += 1
        row += 1

    # Scenario inputs
    for i, res in enumerate(results, 1):
        ws.merge_cells(f"A{row}:D{row}")
        ws.cell(row=row, column=1,
                value=f"Scenario {i}: {res.get('name', 'Unnamed')} ({res.get('type', '')})").font = Font(
            name="Calibri", size=11, bold=True, color="2E86C1"
        )
        row += 1

        inputs = res.get("inputs", {})
        if inputs:
            for k, v in inputs.items():
                ws.cell(row=row, column=1, value=str(k)).font = Font(name="Calibri", size=10, bold=True)
                ws.cell(row=row, column=2, value=str(v)).font = _DATA_FONT
                row += 1
        else:
            ws.cell(row=row, column=1, value="No input data recorded").font = _DATA_FONT
            row += 1
        row += 1

    _auto_width(ws)


def _build_results_sheet(
    wb: Workbook,
    results: List[Dict[str, Any]],
) -> None:
    """Sheet 3: Detailed Results Tables."""
    ws = wb.create_sheet("Results Tables")

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Detailed Results"
    ws["A1"].font = _TITLE_FONT

    row = 3

    for i, res in enumerate(results, 1):
        detail_headers = res.get("table_headers", [])
        detail_rows = res.get("table_rows", [])

        if not detail_headers and not detail_rows:
            continue

        ws.merge_cells(f"A{row}:F{row}")
        ws.cell(row=row, column=1,
                value=f"Scenario {i}: {res.get('name', 'Unnamed')}").font = Font(
            name="Calibri", size=11, bold=True, color="2E86C1"
        )
        row += 1

        if detail_headers:
            for c, h in enumerate(detail_headers, 1):
                ws.cell(row=row, column=c, value=str(h))
            _apply_header_style(ws, row, len(detail_headers))
            row += 1

            for dr in detail_rows:
                for c, val in enumerate(dr, 1):
                    ws.cell(row=row, column=c, value=val)
                row += 1

            _apply_data_style(ws, row - len(detail_rows), row - 1, len(detail_headers))
        row += 1

    _auto_width(ws)


def _build_risk_sheet(
    wb: Workbook,
    results: List[Dict[str, Any]],
) -> None:
    """Sheet 4: Risk Results."""
    ws = wb.create_sheet("Risk Results")

    ws.merge_cells("A1:E1")
    ws["A1"].value = "Quantitative Risk Assessment Results"
    ws["A1"].font = _TITLE_FONT

    row = 3

    for i, res in enumerate(results, 1):
        has_risk_data = any(k in res for k in ("ir_thresholds", "fn_data", "risk_matrix", "risk_level"))

        if not has_risk_data:
            continue

        ws.merge_cells(f"A{row}:E{row}")
        ws.cell(row=row, column=1,
                value=f"Case {i}: {res.get('name', 'Unnamed')}").font = Font(
            name="Calibri", size=11, bold=True, color="2E86C1"
        )
        row += 1

        # IR thresholds
        ir_thresholds = res.get("ir_thresholds", {})
        if ir_thresholds:
            ws.cell(row=row, column=1, value="Individual Risk Contours").font = Font(
                name="Calibri", size=10, bold=True
            )
            row += 1
            ws.cell(row=row, column=1, value="Risk Level")
            ws.cell(row=row, column=2, value="Distance (m)")
            _apply_header_style(ws, row, 2)
            row += 1
            for label, dist in ir_thresholds.items():
                ws.cell(row=row, column=1, value=str(label))
                ws.cell(row=row, column=2, value=dist if isinstance(dist, (int, float)) else str(dist))
                row += 1
            row += 1

        # FN data
        fn_data = res.get("fn_data")
        if fn_data:
            ws.cell(row=row, column=1, value="Societal Risk (FN Curve)").font = Font(
                name="Calibri", size=10, bold=True
            )
            row += 1
            ws.cell(row=row, column=1, value="N (Fatalities)")
            ws.cell(row=row, column=2, value="F (Frequency/year)")
            _apply_header_style(ws, row, 2)
            row += 1

            n_vals = fn_data.get("n", []) if isinstance(fn_data, dict) else []
            f_vals = fn_data.get("f", []) if isinstance(fn_data, dict) else []
            # Handle list format too
            if isinstance(fn_data, list):
                for entry in fn_data:
                    if isinstance(entry, (list, tuple)) and len(entry) >= 2:
                        ws.cell(row=row, column=1, value=entry[0])
                        ws.cell(row=row, column=2, value=entry[1])
                        row += 1
            else:
                for n_val, f_val in zip(n_vals, f_vals):
                    ws.cell(row=row, column=1, value=n_val)
                    ws.cell(row=row, column=2, value=f_val)
                    row += 1
            row += 1

        # Risk level / matrix
        risk_level = res.get("risk_level")
        if risk_level:
            ws.cell(row=row, column=1, value="Risk Level:").font = Font(name="Calibri", size=10, bold=True)
            ws.cell(row=row, column=2, value=str(risk_level)).font = _DATA_FONT
            row += 2

    _auto_width(ws)


def _build_comparison_sheet(
    wb: Workbook,
    results: List[Dict[str, Any]],
) -> None:
    """Sheet 5: Scenario Comparison."""
    ws = wb.create_sheet("Comparison")
    if not results:
        return

    ws.merge_cells("A1:G1")
    ws["A1"].value = "Scenario Comparison"
    ws["A1"].font = _TITLE_FONT

    # Gather all unique summary keys
    all_keys = []
    for res in results:
        summary = res.get("summary", {})
        for key in summary:
            if key not in all_keys:
                all_keys.append(key)

    if not all_keys:
        all_keys = ["max_value"]

    # Headers
    ws.cell(row=3, column=1, value="Parameter")
    for c, res in enumerate(results, 2):
        ws.cell(row=3, column=c, value=res.get("name", f"Scenario {c-1}"))
    _apply_header_style(ws, 3, 1 + len(results))

    row = 4
    for key in all_keys:
        ws.cell(row=row, column=1, value=str(key)).font = Font(name="Calibri", size=10, bold=True)
        for c, res in enumerate(results, 2):
            summary = res.get("summary", {})
            ws.cell(row=row, column=c, value=str(summary.get(key, "")))
        row += 1

    _apply_data_style(ws, 4, row - 1, 1 + len(results))
    _auto_width(ws)


# ══════════════════════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════════════════════

def _extract_substance(result: Dict[str, Any]) -> str:
    """Try to extract substance name from result."""
    inputs = result.get("inputs", {})
    return str(inputs.get("substance", inputs.get("fuel", inputs.get("name", ""))))
